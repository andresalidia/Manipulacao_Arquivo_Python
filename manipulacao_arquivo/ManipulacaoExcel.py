from openpyxl import Workbook, load_workbook

# Criando Planilha 

planilha = Workbook()
tabela = planilha.active
tabela.title = "Planilha Exemplo"
tabela.append(['Nome', 'Idade', 'Cidade'])
tabela.append(['João', 30, 'São Paulo'])
tabela.append(['Maria', 25, 'Rio de Janeiro'])
planilha.save('PrimeiraPlanilhaExcelPython.xlsx')


#ler Planilha

planilha = load_workbook('PrimeiraPlanilhaExcelPython.xlsx')

for row in tabela.iter_rows(values_only=True):
 print(row)

 # editar Planilha

tabela['A2'] = 'José' # Alterar o valor da célula A2
planilha.save('PrimeiraPlanilhaExcelPython.xlsx')

# Mostrar tabela alterada
print("Tabela alterada:")

for row in tabela.iter_rows(values_only=True):
 print(row)

 